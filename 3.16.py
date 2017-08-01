#coding:UTF-8
import openpyxl  
from  openpyxl.workbook  import  Workbook 
from  openpyxl.writer.excel  import  ExcelWriter  
import random
import sys  
reload(sys)  
sys.setdefaultencoding('utf8') #编码问题用3-5行代码解决  
action = raw_input("1.只默写新背内容  2.复习就内容并且默写新内容")
if action == "1":
	num = input("你需要默写多少个单词？") 
	path = raw_input("文件地址：")
	wb = openpyxl.load_workbook(path)  

	#获取workbook中所有的表格  
	sheets = wb.get_sheet_names()  
	print sheets
	# print wb.get_named_ranges()  

	for i in range(len(sheets)):  
	    sheet= wb.get_sheet_by_name(sheets[i])  
	      
	    print('\n\n'+str(i+1)+'sheet: ' + sheet.title+'->>>')  
	    data_list = []
	    for rownum in random.sample(range(1,100),num):
	    	for columnnum in range(1,sheet.max_column+1):
	    		data_list.append(sheet.cell(row = rownum, column = columnnum).value)

	for i in data_list:
		print i
	# 分离英文
	a = []
	for i in range(len(data_list)):
		if i % 2 == 0:
			a.append(data_list[i])
	for i in a:
		print i 

	wb1=Workbook()#新建工作簿
	# ws1 = wb.get_active_sheet()
	# print ws1.title
	# ws1.title = 'New Title'  # 设置worksheet的标题
	new_ws = wb1.create_sheet(title='test')
	key_ws = wb1.create_sheet(title = "test_key")
	i=1
	k = 1  
	for data_l in range(len(data_list)):
		key_ws.cell(column = k , row = i,value = data_list[data_l])
		# new_ws.cell(column = 2 , row = i,value = data_list[data_l])
		if k%2 == 1:
			k = 2
		else:
			k = 1
			i +=1
	i = 1
	for data_m in range(len(a)):
		new_ws.cell(column = 1, row = i, value = a[data_m])
		i +=1
	wb1.save(filename= 'new_file.xlsx')#保存一定要有，否则不会有结果  
elif action == "2":
	num = input("你需要默写多少个单词？")
	num1 = input("你需要复习多少个单词？") 
	path = raw_input("默写文件地址：")
	path1 = raw_input("复习文件地址：")
	wb = openpyxl.load_workbook(path)
	wb_review = openpyxl.load_workbook(path1)  

	#获取workbook中所有的表格  
	sheets = wb.get_sheet_names()
	sheets_review = wb_review.get_sheet_names()  
	print sheets
	# print wb.get_named_ranges()  
	data_list = []
	for i in range(len(sheets)):  
	    sheet= wb.get_sheet_by_name(sheets[i])  

	for i in range(len(sheets_review)):  
	    sheet_review= wb_review.get_sheet_by_name(sheets_review[i])  
	      
	#     print('\n\n'+str(i+1)+'sheet: ' + sheet.title+'->>>')  
   
    for rownum in random.sample(range(1,100),num):
    	for columnnum in range(1,sheet.max_column+1):
    		for rownum1 in random.sample(range(1,100),num1):
    			for columnnum1 in range(1,sheet_review.max_column+1):
    				data_list.append(sheet.cell(row = rownum, column = columnnum).value)
    				data_list.append(sheet_review.cell(row = rownum1, column = columnnum1).value)


	for i in data_list:
		print i
	# 分离英文
	a = []
	for i in range(len(data_list)):
		if i % 2 == 0:
			a.append(data_list[i])
	for i in a:
		print i 

	wb1=Workbook()#新建工作簿
	# ws1 = wb.get_active_sheet()
	# print ws1.title
	# ws1.title = 'New Title'  # 设置worksheet的标题
	new_ws = wb1.create_sheet(title='test')
	key_ws = wb1.create_sheet(title = "test_key")
	i=1
	k = 1  
	for data_l in range(len(data_list)):
		key_ws.cell(column = k , row = i,value = data_list[data_l])
		# new_ws.cell(column = 2 , row = i,value = data_list[data_l])
		if k%2 == 1:
			k = 2
		else:
			k = 1
			i +=1
	i = 1
	for data_m in range(len(a)):
		new_ws.cell(column = 1, row = i, value = a[data_m])
		i +=1
	wb1.save(filename= 'new_file.xlsx')#保存一定要有，否则不会有结果  
